from cgf.config import *
from bdbaza import sqlite as k
from aiogram import types
from aiogram.types import InlineKeyboardMarkup,InlineKeyboardButton, ReplyKeyboardMarkup,KeyboardButton
from bdbaza import sqlite as k
from aiogram.utils.exceptions import BotBlocked
import openpyxl

#глобальные переменные
Text='0'  #глобальная переменная текст пользователя
keksik="""
<b>Жирный шрифт</b>
<code>Текст для копирования</code>
<u>Подчеркнутый текст</u>
<s>Зачеркнутый текст</s>
<a href='vk.com'>Текст ссылки</a>
<i>Курсивный текст</i>"""
############кнопки№############################


def admin_start(): #кнопки стартового меню админки
    markup=ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('Рассылка сообщений пользователям'))
    markup.add(KeyboardButton('Загрузка через EXCEL под КЛЮЧ'))
    markup.add(KeyboardButton('Загрузка через EXCEL время записи'))
    markup.add(KeyboardButton('Загрузка через EXCEL разделов и услуг'))
    markup.add(KeyboardButton('Настройка стартового меню'))
    markup.add(KeyboardButton('Настройка кнопки "Информация"'))
    markup.add(KeyboardButton('Ручное управление'))
    return markup

def standart(): #общие кнопки
    markup=ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('/admin Вернуться в меню'))
    markup.add(KeyboardButton('/start Перезапустить бота'))
    return markup

def proverka_senda(): #кнопки да или нет
    markup=ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('Отправить'))
    markup.insert(KeyboardButton('Перезаписать'))
    markup.add(KeyboardButton('/admin Вернуться в меню'))
    markup.add(KeyboardButton('/start Перезапустить бота'))
    return markup

def knopki_danet():
    markup=ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton('Отправить'))
    markup.insert(KeyboardButton('Перезаписать'))
    markup.add(KeyboardButton('/admin Вернуться в меню'))
    markup.add(KeyboardButton('/start Перезапустить бота'))
    return markup

#########################функции рассылки сообщений###################


async def admin(message:types.Message,state: FSMContext):
    result = await k.take_zapisanie() #
    current_state = await state.get_state()  # текущее машинное состояние пользователя
    if current_state !='':
        await state.finish()
        if admin_key== str(message.chat.id) or admin_key2== str(message.chat.id):
            await message.answer('Здравствуйте, администратор.',reply_markup=admin_start())

async def send_all_message(message:types.Message):
    if admin_key== str(message.chat.id) or admin_key2== str(message.chat.id):
        global keksik

        await message.answer(f'Отправь текст для рассылки всем пользователям в базе:\n'
                             f'{keksik}\n',parse_mode=types.ParseMode.HTML,disable_web_page_preview=True)
        await message.answer(f'Теги для форматирования текста:\n'
                             f'{keksik}\n',reply_markup=standart())
        await Adminka.proverka.set()

async def proverka_text(message:types.Message,state: FSMContext):
    global Text
    Text=message.text
    await message.answer(f'Таким текст отобразиться у пользователей:\n'
                         f' \n'
                         f'{Text}',parse_mode=types.ParseMode.HTML,disable_web_page_preview=True,reply_markup=proverka_senda())
    await Adminka.send.set()

async def final_send_all_message(message:types.Message,state: FSMContext): #массовая рассылка сообщения пользователям
        kek=message.text #проверка на да или нет
        if kek=='Отправить':
            result=await k.take_users_id()
            l=0
            try:
                for i in result:
                        try:
                            await bot.send_message(i[0],Text,parse_mode=types.ParseMode.HTML,disable_web_page_preview=True)#последний параметр чтобы нельзя было предпросмотреть ссылку
                            l+=1
                            print(f'Отправлено {l} пользователям')
                        except BotBlocked as exception_info:
                            print(f'Юзер с айди {i[0]} заблокировал чат')
                            print(exception_info)
            except Exception:
                await message.answer('Произошла ошибка, попробуйте ввести текст снова:')
                print('Произошла ошибка вероятно с HTML')
            await message.answer(f'Разослано {l} пользователям',reply_markup=standart())
            await state.finish()
        elif kek=='Перезаписать':
            global keksik
            await message.answer(f'Отправь текст для рассылки всем пользователям в базе:\n'
                                 f'{keksik}\n', parse_mode=types.ParseMode.HTML, disable_web_page_preview=True)
            await message.answer(f'Теги для форматирования текста:\n'
                                 f'{keksik}\n', reply_markup=standart())
            await Adminka.proverka.set()
        else:
            print('Что то другое')

########################################Загрузка дат и окошей через excel############################################

async def start_excel(message:types.Message,state:FSMContext):
    if admin_key== str(message.chat.id) or admin_key2== str(message.chat.id):
        await message.answer_document(open('cgf/shablon.xlsx', 'rb'))
        await message.answer('Вы можете добавить время с датами через EXCEL файл \n'
                             'Шаблон для заполнения данных <b>выше</b>⬆⬆⬆\n'
                             'Чтобы продолжить,<b> отправьте заполненный шаблон </b> в ответном сообщении или нажмите /admin\n'
                             '<b>ИЗ ЭТОГО МЕНЮ, ВЫ СМОЖЕТЕ ЗАГРУЗИТЬ ТОЛЬКО ДАТУ И ВРЕМЯ</b>',parse_mode=types.ParseMode.HTML,reply_markup=standart())
        await Adminka.take_excel.set()

async def take_excel(message: types.Message):  # загрузка excel файла в
    try:
        result = await k.take_users_id()
        print("downloading document")
        #destination = r"cgf\file1.xlsx"
        await message.document.download(destination_file = r"cgf/file1.xlsx")
        book = openpyxl.load_workbook(r"cgf/file1.xlsx", read_only=True)
        sheet = book.active  # активирует листы
        if sheet[1][0].value=='Даты':
            await message.answer('Файл загружен, начинаю обработку:')
            await zapis_iz_excel(message)
        else:

            await message.answer('Ошибка загрузки файла(возможно вы не загружаете шаблон')
    except:
        await message.answer('Произошла ошибка загрузки шаблона,проверьте заполнение',reply_markup=standart())

async def zapis_iz_excel(message): #метод чтения данных из екселя и записей в бд вызывается в take_excel
    await k.delet_ne_zapisanie()#удаление при загрузке из excel всех окошек
    result = await k.take_zapisanie() #получение дат на которые уже есть записи
    book=openpyxl.load_workbook("cgf/file1.xlsx",read_only=True)
    sheet = book.active #активирует листы
    print(sheet[1][0].value) #1 = это строка #2 - это столбец

    for row in range(2,33): #от 1 строки до 33 в екселе
        date = sheet[row][0].value #вывести первый столбец дата
        times = sheet[row][1].value #вывести второй столбец время
        if sheet[row][0].value==None:
            await message.answer(f'Загружено до {sheet[row-1][0].value}')
            break
        else:
            dates=date.strftime("%Y-%m-%d")
            time_all=times.split(',') #формат окошек
            for dat in result:  # проверяем на записи на дни которые уже записанные №000
                if dat[0]==dates: # сравниваем даты которые не удалились(на которые записанные_ с датой из екселя
                    await message.answer(f'На {dates} имеется запись, перезапись невозможно. Удалите в ручную')
                    break #останавливаем запись
            else: #продолжаем запись
                for i in time_all:
                    await k.do_date(dates,i)

    await message.answer('Загрузка дат выполнена, вернитесь в главное меню',reply_markup=standart())


########################################Загрузка разделов и подразделов############################################

async def zapis_razdelov_uslug(message: types.Message):
    if admin_key== str(message.chat.id) or admin_key2== str(message.chat.id):
        await message.answer_document(open('cgf/shablon.xlsx', 'rb'))
        await message.answer('Вы можете добавить разделы и услуги через EXCEL файл \n'
                             'Шаблон для заполнения данных <b>выше</b>⬆⬆⬆\n'
                             'Чтобы продолжить,<b> отправьте заполненный шаблон </b> в ответном сообщении или нажмите /admin\n'
                             '<b>ИЗ ЭТОГО МЕНЮ, ВЫ СМОЖЕТЕ ЗАГРУЗИТЬ ТОЛЬКО РАЗДЕЛЫ И УСЛУГИ</b>',parse_mode=types.ParseMode.HTML,reply_markup=standart())
        await Adminka.take_excel_uslugi.set()


async def take_excel_uslugi(message: types.Message):  # загрузка excel файла в
    #try:
        result = await k.take_users_id()
        print("downloading document")
        #destination = r"cgf\file1.xlsx"
        await message.document.download(destination_file = r"cgf/file1.xlsx")
        book = openpyxl.load_workbook(r"cgf/file1.xlsx", read_only=True)
        sheet = book.active  # активирует листы
        if sheet[1][0].value=='Даты':
            await message.answer('Файл загружен, начинаю обработку:')
            await zapis_iz_excel_uslugi(message) #вызов функции обработки
        else:
            await message.answer('Ошибка загрузки файла(возможно вы не отправляйте шаблон')

async def zapis_iz_excel_uslugi(message): #метод чтения данных из екселя и записей в бд вызывается в take_excel
    try:
        book=openpyxl.load_workbook("cgf/file1.xlsx",read_only=True)
        sheet = book.active #активирует листы
        await k.delet_uslugi()
        for row in range(3,9): #от 1 строки до 33 в екселе
            usluga = sheet[row][4].value #вывести первый столбец дата
            if usluga==None:
                await message.answer(f'Последняя запись: {sheet[row-1][4].value}')
                break
            else:
                await k.insert_uslugi(usluga)
        await message.answer('Запись разделов услуг выполнена, перехожу к обработке услуг')
    except:
        await message.answer('Возникла ошибка при попытке записи разделов услуг')
    try:
        for row in range(3,79):
            us = sheet[row][6].value
            if us == None:
                await message.answer(f'Последняя запись: {sheet[row - 1][6].value}')
                break
            else:
                us=sheet[row][6].value.split(':')
                await k.insert_alluslugi(us[0],us[1],us[2])
        await message.answer('Выполнено успешно!',reply_markup=standart())
    except:
        await message.answer('Возникла ошибка при попытке записи самих услуг')

##############загрузка всего из шаблона под КЛЮЧ #############################################

async def all_excel(message: types.Message): #запись под ключ всего файла из екселя сразу
    if admin_key== str(message.chat.id) or admin_key2== str(message.chat.id):
        await message.answer_document(open('cgf/shablon.xlsx', 'rb'))
        await message.answer('Вы можете за<b> один раз </b> загрузить все данные через EXCEL файл \n'
                             'Шаблон для заполнения данных <b>выше</b>⬆⬆⬆\n'
                             'Чтобы продолжить,<b> отправьте заполненный шаблон </b> в ответном сообщении или нажмите /admin\n'
                             '<b>ПЕРЕЗАПИСЬ ВСЕХ ДАННЫХ(дат,времени, разделов, услуг через данный шаблон</b>',parse_mode=types.ParseMode.HTML,reply_markup=standart())
        await Adminka.final_excel.set()

async def all_excel_final(message:types.Message):
    print('метооод')
    try:
        result = await k.take_users_id()
        print("downloading document")
        await message.document.download(destination_file=r"cgf/file1.xlsx")
        book = openpyxl.load_workbook(r"cgf/file1.xlsx", read_only=True)
        sheet = book.active  # активирует листы
        if sheet[1][0].value == 'Даты':
            await message.answer('Файл загружен, начинаю обработку:')
            await zapis_iz_excel(message)
            await zapis_iz_excel_uslugi(message)
        else:
            await message.answer('Ошибка загрузки файла(возможно вы не отправляйте шаблон')
    except:
        await message.answer('Произошла ошибка в загрузке ПОД КЛЮЧ')






#регистрация хендлеров#
def register_admin_handler(dp: Dispatcher) -> None:  # функция регистрации в дальнейшем передадим ее в мейн, и здесь
    #хендлеры сообщениииий
    dp.register_message_handler(admin, commands=['admin'],state='*')  # старт
    dp.register_message_handler(send_all_message, text='Рассылка сообщений пользователям')  # старт
    dp.register_message_handler(proverka_text,state=Adminka.proverka)  # текст на проверку отправить да нет
    dp.register_message_handler(final_send_all_message,state=Adminka.send)  # старт
    #хендлеры с excel
    dp.register_message_handler(start_excel,text='Загрузка через EXCEL время записи')
    dp.register_message_handler(take_excel,content_types=types.ContentType.DOCUMENT,state=Adminka.take_excel)
    dp.register_message_handler(zapis_razdelov_uslug,text='Загрузка через EXCEL разделов и услуг')
    dp.register_message_handler(take_excel_uslugi,content_types=types.ContentType.DOCUMENT,state=Adminka.take_excel_uslugi)
    dp.register_message_handler(all_excel,text='Загрузка через EXCEL под КЛЮЧ')
    dp.register_message_handler(all_excel_final,content_types=types.ContentType.DOCUMENT,state=Adminka.final_excel)